from django.views.decorators.http import require_GET
import json
from urllib.parse import urlencode
from django.shortcuts import redirect, get_object_or_404, render
from django.http import HttpResponseForbidden, HttpResponse
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.csrf import csrf_exempt, ensure_csrf_cookie
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework.response import Response
from rest_framework.views import APIView

from rest_framework_simplejwt.authentication import JWTAuthentication
from accounts.models import User
from .models import Incident
from .services import assign_nearest_online_officer, assign_incident_to_officer


# =========================
# Helpers
# =========================
def is_admin(user):
    return user.is_authenticated and user.is_staff


def status_uz(s: str) -> str:
    mp = {
        "NEW": "Yangi",
        "ASSIGNED": "Biriktirildi",
        "ACCEPTED": "Qabul qilindi",
        "ARRIVED": "Yetib keldi",
        "DONE": "Bajarildi",
        "CLOSED": "Yopildi",
        "CANCELED": "Bekor qilindi",
    }
    return mp.get(s, s or "")


# =========================
# (Ixtiyoriy) Admin dashboard view (map.html)
# Agar sende admin_dashboard tracking.views da bo'lsa, BU FUNKSIYANI O'CHIR.
# =========================
@login_required
@user_passes_test(is_admin)
def admin_dashboard(request):
    group = request.GET.get("group", "PROF")
    return render(request, "dashboard/map.html", {"group": group})


# =========================
# Dashboard form: Incident yaratish (Admin dashboarddagi forma)
# =========================
@csrf_exempt
def create_incident_form(request):
    if request.method != "POST":
        return redirect("/dashboard/")

    lat = request.POST.get("lat")
    lng = request.POST.get("lng")

    if lat and lng:
        incident = Incident.objects.create(
            title="Vizov",
            lat=float(lat),
            lng=float(lng),
            address_text=request.POST.get("address_text", ""),
            details=request.POST.get("details", ""),
            victim_first_name=request.POST.get("victim_first_name", ""),
            victim_last_name=request.POST.get("victim_last_name", ""),
            caller_phone=request.POST.get("caller_phone", ""),
            status="NEW",
        )
        assign_nearest_online_officer(incident)

    return redirect("/dashboard/")


# =========================
# API: Incident yaratish (Admin/Operator) - JSON
# =========================
class CreateIncidentView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]

    def post(self, request):
        lat = request.data.get("latitude")
        lng = request.data.get("longitude")

        if lat is None or lng is None:
            return Response({"detail": "latitude and longitude required"}, status=400)

        incident = Incident.objects.create(
            title=request.data.get("title", "Vizov"),
            details=request.data.get("details", ""),
            address_text=request.data.get("address_text", ""),
            victim_first_name=request.data.get("victim_first_name", ""),
            victim_last_name=request.data.get("victim_last_name", ""),
            caller_phone=request.data.get("caller_phone", ""),
            lat=float(lat),
            lng=float(lng),
            status="NEW",
            created_by=request.user,
        )

        assign_nearest_online_officer(incident)
        return Response({"ok": True, "id": incident.id})

class AssignIncidentView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]

    def post(self, request, pk):
        inc = get_object_or_404(Incident, pk=pk)

        officer_id = request.data.get("officer_id")
        if not officer_id:
            return Response({"detail": "officer_id required"}, status=400)

        officer = get_object_or_404(User, pk=officer_id)

        if getattr(officer, "role", None) != User.Role.OFFICER:
            return Response({"detail": "Selected user is not OFFICER"}, status=400)

        assign_incident_to_officer(inc, officer)

        return Response({
            "ok": True,
            "incident_id": inc.id,
            "assigned_officer_id": officer.id,
            "assigned_officer_username": officer.username,
            "status": inc.status,
        })
# =========================
# API: Incidentlar ro'yxati (Admin uchun FULL)
# =========================
class IncidentListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = Incident.objects.select_related("assigned_officer").order_by("-id")

        data = []
        for inc in qs:
            officer_username = None
            officer_name = None

            if getattr(inc, "assigned_officer", None):
                officer_username = getattr(inc.assigned_officer, "username", None)
                full_name = ""
                if hasattr(inc.assigned_officer, "get_full_name"):
                    full_name = (inc.assigned_officer.get_full_name() or "").strip()
                officer_name = full_name or officer_username

            data.append({
                "id": inc.id,
                "status": inc.status,
                "lat": getattr(inc, "lat", None),
                "lng": getattr(inc, "lng", None),
                "created_at": inc.created_at.isoformat() if getattr(inc, "created_at", None) else None,
                "accepted_at": inc.accepted_at.isoformat() if getattr(inc, "accepted_at", None) else None,
                "arrived_at": inc.arrived_at.isoformat() if getattr(inc, "arrived_at", None) else None,
                "closed_at": inc.closed_at.isoformat() if getattr(inc, "closed_at", None) else None,
                "details": getattr(inc, "details", "") or "",
                "description": getattr(inc, "description", "") or "",
                "address_text": getattr(inc, "address_text", "") or "",
                "address": getattr(inc, "address", "") or "",
                "caller_phone": getattr(inc, "caller_phone", "") or "",
                "phone": getattr(inc, "phone", "") or "",
                "assigned_officer": inc.assigned_officer_id,
                "assigned_officer_username": officer_username,
                "assigned_officer_name": officer_name,
            })

        return Response(data)


class OfficerMyIncidentsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if getattr(request.user, "role", None) != User.Role.OFFICER:
            return Response({"detail": "Only OFFICER can view this."}, status=403)

        qs = Incident.objects.select_related("assigned_officer").filter(
            assigned_officer_id=request.user.id,
            status__in=["ASSIGNED", "ACCEPTED", "ARRIVED"]
        ).order_by("-id")

        data = []
        for inc in qs:
            officer_username = None
            officer_name = None

            if getattr(inc, "assigned_officer", None):
                officer_username = getattr(inc.assigned_officer, "username", None)
                full_name = ""
                if hasattr(inc.assigned_officer, "get_full_name"):
                    full_name = (inc.assigned_officer.get_full_name() or "").strip()
                officer_name = full_name or officer_username

            data.append({
                "id": inc.id,
                "status": inc.status,
                "lat": getattr(inc, "lat", None),
                "lng": getattr(inc, "lng", None),
                "created_at": inc.created_at.isoformat() if getattr(inc, "created_at", None) else None,
                "accepted_at": inc.accepted_at.isoformat() if getattr(inc, "accepted_at", None) else None,
                "arrived_at": inc.arrived_at.isoformat() if getattr(inc, "arrived_at", None) else None,
                "closed_at": inc.closed_at.isoformat() if getattr(inc, "closed_at", None) else None,
                "details": getattr(inc, "details", "") or "",
                "description": getattr(inc, "description", "") or "",
                "address_text": getattr(inc, "address_text", "") or "",
                "address": getattr(inc, "address", "") or "",
                "caller_phone": getattr(inc, "caller_phone", "") or "",
                "phone": getattr(inc, "phone", "") or "",
                "assigned_officer": inc.assigned_officer_id,
                "assigned_officer_username": officer_username,
                "assigned_officer_name": officer_name,
            })

        return Response(data)

# =========================
# Dashboard: Officer panel
# =========================
@ensure_csrf_cookie
@login_required
def officer_dashboard(request):
    if getattr(request.user, "role", None) != User.Role.OFFICER:
        return HttpResponseForbidden("Only OFFICER can access.")

    incidents = Incident.objects.filter(
        assigned_officer=request.user
    ).order_by("-created_at")

    return render(request, "dashboard/officer.html", {
        "me": request.user,
        "incidents": incidents,
    })


# =========================
# API: Officer status update (mobil app uchun ham)
# JSON: {"status":"ACCEPTED"}
# =========================
class OfficerIncidentStatusView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        if getattr(request.user, "role", None) != User.Role.OFFICER:
            return Response({"detail": "Only OFFICER can do this."}, status=403)

        inc = get_object_or_404(Incident, pk=pk)

        if inc.assigned_officer_id != request.user.id:
            return Response({"detail": "Not your incident."}, status=403)

        new_status = request.data.get("status")
        if new_status not in ["ACCEPTED", "ARRIVED", "DONE", "CLOSED", "CANCELED"]:
            return Response({"detail": "Invalid status"}, status=400)

        now = timezone.now()

        if new_status == "ACCEPTED" and not inc.accepted_at:
            inc.accepted_at = now
        if new_status == "ARRIVED" and not inc.arrived_at:
            inc.arrived_at = now
        if new_status in ["DONE", "CLOSED"] and not inc.closed_at:
            inc.closed_at = now

        inc.status = new_status
        inc.save()

        return Response({"ok": True, "status": inc.status})


# =========================
# REPORT: daily Excel
# URL: /reports/daily.xlsx   (ixtiyoriy: ?date=2026-02-15)
# =========================
@login_required
@user_passes_test(is_admin)
def daily_report_xlsx(request):
    date_str = request.GET.get("date")
    today = timezone.localdate()

    if date_str:
        try:
            y, m, d = [int(x) for x in date_str.split("-")]
            day = timezone.datetime(y, m, d).date()
        except Exception:
            day = today
    else:
        day = today

    qs = Incident.objects.select_related("assigned_officer").filter(
        created_at__date=day
    ).order_by("id")

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily"

    headers = [
        "Sana",
        "Vizov ID",
        "Murojaatchi ismi",
        "Murojaatchi familiya",
        "Tel",
        "Hodisa tafsiloti",
        "Manzil",
        "Hodim",
        "Qabul vaqti",
        "Yetib kelish vaqti",
        "Yopilgan vaqti",
        "Yetib borish (min)",
        "Yopilishgacha davomiylik (min)",
        "Status",
    ]
    ws.append(headers)

    def mins_between(a, b):
        if not a or not b:
            return ""
        return int((b - a).total_seconds() // 60)

    for inc in qs:
        accepted = getattr(inc, "accepted_at", None)
        arrived = getattr(inc, "arrived_at", None)
        closed = getattr(inc, "closed_at", None)

        officer_name = ""
        if inc.assigned_officer:
            officer_name = (
                inc.assigned_officer.get_full_name().strip()
                if hasattr(inc.assigned_officer, "get_full_name") and inc.assigned_officer.get_full_name().strip()
                else getattr(inc.assigned_officer, "username", str(inc.assigned_officer))
            )

        ws.append([
            str(day),
            inc.id,
            getattr(inc, "victim_first_name", "") or "",
            getattr(inc, "victim_last_name", "") or "",
            getattr(inc, "caller_phone", "") or "",
            getattr(inc, "details", "") or "",
            getattr(inc, "address_text", "") or "",
            officer_name,
            accepted.strftime("%d.%m.%Y %H:%M:%S") if accepted else "",
            arrived.strftime("%d.%m.%Y %H:%M:%S") if arrived else "",
            closed.strftime("%d.%m.%Y %H:%M:%S") if closed else "",
            mins_between(accepted, arrived),
            mins_between(accepted, closed),
            status_uz(getattr(inc, "status", "")),
        ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="daily_{day}.xlsx"'
    wb.save(resp)
    return resp


# =========================
# DETAIL PAGE: /dashboard/incidents/<id>/
# =========================
@login_required
@user_passes_test(is_admin)
def incident_detail_page(request, pk):
    inc = get_object_or_404(
        Incident.objects.select_related("assigned_officer"),
        pk=pk
    )
    return render(request, "dashboard/incident_detail.html", {"inc": inc})


@csrf_exempt
def officer_location(request):
    if request.method != "POST":
        return JsonResponse({"error": "POST only"}, status=405)

    try:
        auth = JWTAuthentication()
        auth_result = auth.authenticate(request)

        if auth_result is None:
            return JsonResponse({"error": "Token invalid or missing"}, status=401)

        user, token = auth_result

        raw = request.body.decode("utf-8")
        print("RAW BODY:", raw)

        data = json.loads(raw)

        latitude = data.get("latitude")
        longitude = data.get("longitude")
        accuracy = data.get("accuracy")

        print("USER:", user)
        print("OFFICER LOCATION:", latitude, longitude, accuracy)

        model_fields = {f.name for f in user._meta.fields}
        update_fields = []

        if "latitude" in model_fields:
            user.latitude = latitude
            update_fields.append("latitude")

        if "longitude" in model_fields:
            user.longitude = longitude
            update_fields.append("longitude")

        if "accuracy" in model_fields:
            user.accuracy = accuracy
            update_fields.append("accuracy")

        if "last_seen" in model_fields:
            user.last_seen = timezone.now()
            update_fields.append("last_seen")

        if "is_online" in model_fields:
            user.is_online = True
            update_fields.append("is_online")

        if update_fields:
            user.save(update_fields=update_fields)

        return JsonResponse({
            "ok": True,
            "latitude": latitude,
            "longitude": longitude,
            "accuracy": accuracy,
            "updated_fields": update_fields
        })

    except Exception as e:
        print("LOCATION ERROR:", str(e))
        return JsonResponse({"error": str(e)}, status=400)


class MyProfileView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        u = request.user

        photo_url = ""
        try:
            if u.photo:
                photo_url = request.build_absolute_uri(u.photo.url)
        except Exception:
            photo_url = ""

        return Response({
            "id": u.id,
            "username": u.username,
            "first_name": u.first_name,
            "last_name": u.last_name,
            "full_name": (u.first_name + " " + u.last_name).strip(),
            "role": u.role,
            "position": u.position,
            "phone": u.phone,
            "photo": photo_url
        })
@require_GET
def search_location(request):
    q = request.GET.get("q", "").strip()

    if not q:
        return JsonResponse([], safe=False)

    params = urlencode({
        "q": q,
        "format": "jsonv2",
        "addressdetails": 1,
        "namedetails": 1,
        "extratags": 1,
        "countrycodes": "uz",
        "limit": 10,
    })

    url = "https://nominatim.openstreetmap.org/search?" + params

    try:
        req = Request(
            url,
            headers={
                "User-Agent": "PatrolDashboard/1.0",
                "Accept": "application/json",
            },
        )

        with urlopen(req, timeout=8) as response:
            data = json.loads(response.read().decode("utf-8"))

        return JsonResponse(data, safe=False)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
